import pandas as pd, tkinter as tk
from tkinter import *
from openpyxl import load_workbook
from tkinter import messagebox
from pandas import Series, DataFrame
import matplotlib.pyplot as plt
from statistics import mean
from PIL import Image, ImageTk
from openpyxl.workbook import Workbook
sales = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=0)
care = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=1)
skyservice = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=2)
sky = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=3)
vmcare = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=4)
vmcablecare = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=5)
LM = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=6)
Verizon = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=7)
phase3_gendtvsales = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=8)
phase3_gendhsclg = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=9)
phase3_gendhssales = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=10)
phase4_dmdr = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=11)
phase4_ismclg = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=12)
phase4_ismsvc = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=13)
phase4_mobclg = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=14)
phase4_mobss = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=15)
nurseline = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=16)
housecalls = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name=17)
wb = Workbook()

def get_factors(df, name):
    writer = pd.ExcelWriter('Test.xlsx', engine='openpyxl', mode='a')
    EBP_UPTIME = round(mean(df['UPTIMEPERCENTAGE'] * 100), 2)
    GAIN_ACHIEVEMENT = round(df['GAINACHIEVEMENT'] * 100, 2)
    GAIN_ACHIEVEMENT = GAIN_ACHIEVEMENT[1]
    REVENUE_ACHIEVEMENT = round(df.iloc[0]['REVENUEACHIEVEMENT'], 2)
    missingvalue(df)
    Oncall = sum(df['ONCALLS'])
    a = [
     'ONCALLS', 'OFFCALLS', 'BENCHMARK', 'SUCCESSROUTES', 'FAILROUTEPERC', 'OFFAGENTSLA', 'OFFAGENTSLA%AGE', 'ONAGENTSLA', 'ONAGENTSLA%AGE', 'OFFCALLSLA', 'OFFCALLSLA%AGE', 'ONCALLSLA', 'ONCALLSLA%AGE', '1-1CALLS', '1-1CALLS%AGE', '1-1CALLSWITHOUTSLABLOWNS', '1-1CALLS%AGEWITHOUTSLABLOWNS', 'L2CALLS', 'L2CALLS%AGE', 'O0BANDONS', 'OFFABANDONS', 'O0BANDONSPERC', 'OFFABANDONSPERC', 'O0BAN-OFFABANPERC', 'O0P', 'OFFAP', 'APSKEW', 'ONCP', 'OFFCP', 'AGENTCHOICE', 'FILTEREDAGENTCHOICE', 'USEDAGENTCHOIDEWITHOUTSLABLOWNS', 'CALLCHOICE', 'FILTEREDCALLCHOICE', 'USEDCALLCHOICEWIHOUTSLABLOWNS', 'ONEVALSCORERAW', 'OFFEVALSCORERAW', 'ONEVALSCOREUSED', 'OFFEVALSCOREUSED', 'ONEVALUATIONERRCALLS', 'ONEVALUATIONERRCALLS%AGE', 'OFFEVALUATIONERRCALLS', 'OFFEVALUATIONERRCALLS%AGE', 'LOOKUPFAILURES', 'LOOKUPFAILUREPERC', 'UNKNOWNAGENTCALLS', 'UNKNOWNAGENTCALLS%AGE', 'CGNOTFOUNDCALLS', 'CGNOTFOUNDCALLS%AGE']
    Actual_Values = {}
    for i in a:
        Actual_Values[i] = round(sum(df[i]) * 100 / Oncall, 2)
    else:
        One_One_Calls_Actual = Actual_Values['1-1CALLS']
        One_One_Calls_Impact = round(df.iloc[0]['1-1CALLSIMPACT'] * 100, 2)
        Unknown_Agent_Calls_Actual = Actual_Values['UNKNOWNAGENTCALLS']
        Unknown_Agent_Calls_Impact = round(df.iloc[0]['UNKNOWNAGENTIMPACT'] * 100, 2)
        data1 = pd.DataFrame(Actual_Values.items())
        data2 = pd.DataFrame([['EBP_UPTIME', 'GAIN_ACHIEVEMENT', 'REVENUE_ACHIEVEMENT', 'One_One_Calls_Actual', 'One_One_Calls_Impact', 'Unknown_Agent_Calls_Actual', 'Unknown_Agent_Calls_Impact'], [EBP_UPTIME, GAIN_ACHIEVEMENT, REVENUE_ACHIEVEMENT, One_One_Calls_Actual, One_One_Calls_Impact, Unknown_Agent_Calls_Actual, Unknown_Agent_Calls_Impact]])
        data1 = data1.append(data2)
        if name == 'Bygs_Sales':
            data1.to_excel(writer, sheet_name='Sales_Calculations')
            writer.save()
            writer.close()
        elif name == 'Bygs_Care':
            data1.to_excel(writer, sheet_name='Care_Calculations')
            writer.save()
            writer.close()


def get_factorsSkyService(df, name):
    writer = pd.ExcelWriter('Test.xlsx', engine='openpyxl', mode='a')
    EBP_UPTIME = round(mean(df['UPTIMEPERCENTAGE'] * 100), 2)
    GAIN_ACHIEVEMENT = round(df['GAINACHIEVEMENT'] * 100, 2)
    GAIN_ACHIEVEMENT = GAIN_ACHIEVEMENT[1]
    REVENUE_ACHIEVEMENT = round(df.iloc[0]['REVENUEACHIEVEMENT'], 2)
    missingvalue(df)
    Oncall = sum(df['ONCALLS'])
    a = [
     'EXPECTGAIN', 'GAIN', 'GAINACHIEVEMENT',
     'REVENUEACHIEVEMENT', 'ANYCRITICALISSUE', 'DOWNTIMEINMINS',
     'REVENUEIMPACT', 'SCHEDULEDTIME', 'UPTIMEPERCENTAGE',
     'FAILEDROUTEIMPACT', 'ONAGENTSLAIMPACT', 'ONCALLSLAIMPACT',
     '1-1CALLSIMPACT', 'ONEVALUATIONERRIMPACT', 'LOOKUPFAILUREIMPACT',
     'UNKNOWNAGENTIMPACT', 'CGNOTFOUNDIMPACT', 'DISTINCTAGENTS',
     'PREVIOUSTOTALCALLS', 'CALLDIFFPERC', 'TOTALCALLS', 'ONCALLS',
     'OFFCALLS', 'BENCHMARK', 'SUCCESSROUTES', 'FAILROUTEPERC',
     'OFFAGENTSLA', 'OFFAGENTSLA%AGE', 'ONAGENTSLA', 'ONAGENTSLA%AGE',
     'OFFCALLSLA', 'OFFCALLSLA%AGE', 'ONCALLSLA', 'ONCALLSLA%AGE',
     '1-1CALLS', '1-1CALLS%AGE', '1-1CALLSWITHOUTSLABLOWNS',
     '1-1CALLS%AGEWITHOUTSLABLOWNS', 'L2CALLS', 'L2CALLS%AGE', 'O0BANDONS',
     'OFFABANDONS', 'O0BANDONSPERC', 'OFFABANDONSPERC', 'O0BAN-OFFABANPERC',
     'O0P', 'OFFAP', 'APSKEW', 'ONCP', 'OFFCP', 'AGENTCHOICE',
     'USEDAGENTCHOICE', 'USEDAGENTCHOICEWITHOUTSLABLOWNS', 'CALLCHOICE',
     'USEDCALLCHOICE', 'USEDCALLCHOICEWITHOUTSLABLOWNS', 'ONEVALSCORERAW',
     'OFFEVALSCORERAW', 'ONEVALSCOREUSED', 'OFFEVALSCOREUSED',
     'ONEVALUATIONERRCALLS', 'ONEVALUATIONERRCALLS%AGE',
     'OFFEVALUATIONERRCALLS', 'OFFEVALUATIONERRCALLS%AGE', 'LOOKUPFAILURES',
     'LOOKUPFAILUREPERC', 'UNKNOWNAGENTCALLS', 'UNKNOWNAGENTCALLS%AGE',
     'CGNOTFOUNDCALLS', 'CGNOTFOUNDCALLS%AGE',
     'IMPORTANT/HIGH/CRITICALTICKETS']
    Actual_Values = {}
    for i in a:
        Actual_Values[i] = round(sum(df[i]) * 100 / Oncall, 2)
    else:
        One_One_Calls_Actual = Actual_Values['1-1CALLS']
        One_One_Calls_Impact = round(df.iloc[0]['1-1CALLSIMPACT'] * 100, 2)
        Unknown_Agent_Calls_Actual = Actual_Values['UNKNOWNAGENTCALLS']
        Unknown_Agent_Calls_Impact = round(df.iloc[0]['UNKNOWNAGENTIMPACT'] * 100, 2)
        data1 = pd.DataFrame(Actual_Values.items())
        data2 = pd.DataFrame([['EBP_UPTIME', 'GAIN_ACHIEVEMENT', 'REVENUE_ACHIEVEMENT', 'One_One_Calls_Actual', 'One_One_Calls_Impact', 'Unknown_Agent_Calls_Actual', 'Unknown_Agent_Calls_Impact'], [EBP_UPTIME, GAIN_ACHIEVEMENT, REVENUE_ACHIEVEMENT, One_One_Calls_Actual, One_One_Calls_Impact, Unknown_Agent_Calls_Actual, Unknown_Agent_Calls_Impact]])
        data1 = data1.append(data2)
        data1.to_excel(writer, sheet_name='SkyService_Calculations')
        writer.save()
        writer.close()


def get_factorsSky(df, name):
    writer = pd.ExcelWriter('Test.xlsx', engine='openpyxl', mode='a')
    EBP_UPTIME = round(mean(df['UPTIMEPERCENTAGE'] * 100), 2)
    GAIN_ACHIEVEMENT = round(df['GAINACHIEVEMENT'] * 100, 2)
    GAIN_ACHIEVEMENT = GAIN_ACHIEVEMENT[1]
    REVENUE_ACHIEVEMENT = round(df.iloc[0]['REVENUEACHIEVEMENT'], 2)
    missingvalue(df)
    Oncall = sum(df['ONCALLS'])
    a = ['EXPECTGAIN', 'GAIN',
     'GAINACHIEVEMENT', 'REVENUEACHIEVEMENT', 'DOWNTIMEINMINS',
     'REVENUEIMPACT', 'DISTINCTAGENTS', 'SCHEDULEDTIME', 'UPTIMEPERCENTAGE',
     'FAILEDROUTEIMPACT', 'ONAGENTSLAIMPACT', 'ONCALLSLAIMPACT',
     '1-1CALLSIMPACT', 'ONEVALUATIONERRIMPACT', 'LOOKUPFAILUREIMPACT',
     'UNKNOWNAGENTIMPACT', 'CGNOTFOUNDIMPACT', 'PREVIOUSTOTALCALLS',
     'CALLDIFFPERC', 'TOTALCALLS', 'ONCALLS', 'OFFCALLS', 'BENCHMARK',
     'SUCCESSROUTES', 'FAILROUTEPERC', 'OFFAGENTSLA', 'OFFAGENTSLA%AGE',
     'ONAGENTSLA', 'ONAGENTSLA%AGE', 'OFFCALLSLA', 'OFFCALLSLA%AGE',
     'ONCALLSLA', 'ONCALLSLA%AGE', '1-1CALLS', '1-1CALLS%AGE',
     '1-1CALLSWITHOUTSLABLOWNS', '1-1CALLS%AGEWITHOUTSLABLOWNS', 'L2CALLS',
     'L2CALLS%AGE', 'O0BANDONS', 'OFFABANDONS', 'O0BANDONSPERC',
     'OFFABANDONSPERC', 'O0BAN-OFFABANPERC', 'O0P', 'OFFAP', 'APSKEW',
     'ONCP', 'OFFCP', 'AGENTCHOICE', 'USEDAGENTCHOICE',
     'USEDAGENTCHOICEWITHOUTSLABLOWNS', 'CALLCHOICE', 'USEDCALLCHOICE',
     'USEDCALLCHOICEWITHOUTSLABLOWNS', 'ONEVALSCORERAW', 'OFFEVALSCORERAW',
     'ONEVALSCOREUSED', 'OFFEVALSCOREUSED', 'ONEVALUATIONERRCALLS',
     'ONEVALUATIONERRCALLS%AGE', 'OFFEVALUATIONERRCALLS',
     'OFFEVALUATIONERRCALLS%AGE', 'LOOKUPFAILURES', 'LOOKUPFAILUREPERC',
     'UNKNOWNAGENTCALLS', 'UNKNOWNAGENTCALLS%AGE', 'CGNOTFOUNDCALLS',
     'CGNOTFOUNDCALLS%AGE', 'IMPORTANT/HIGH/CRITICALTICKETS']
    Actual_Values = {}
    for i in a:
        Actual_Values[i] = round(sum(df[i]) * 100 / Oncall, 2)
    else:
        One_One_Calls_Actual = Actual_Values['1-1CALLS']
        One_One_Calls_Impact = round(df.iloc[0]['1-1CALLSIMPACT'] * 100, 2)
        Unknown_Agent_Calls_Actual = Actual_Values['UNKNOWNAGENTCALLS']
        Unknown_Agent_Calls_Impact = round(df.iloc[0]['UNKNOWNAGENTIMPACT'] * 100, 2)
        data1 = pd.DataFrame(Actual_Values.items())
        data2 = pd.DataFrame([['EBP_UPTIME', 'GAIN_ACHIEVEMENT', 'REVENUE_ACHIEVEMENT', 'One_One_Calls_Actual', 'One_One_Calls_Impact', 'Unknown_Agent_Calls_Actual', 'Unknown_Agent_Calls_Impact'], [EBP_UPTIME, GAIN_ACHIEVEMENT, REVENUE_ACHIEVEMENT, One_One_Calls_Actual, One_One_Calls_Impact, Unknown_Agent_Calls_Actual, Unknown_Agent_Calls_Impact]])
        data1 = data1.append(data2)
        data1.to_excel(writer, sheet_name='Sky_Calculations')
        writer.save()
        writer.close()


def get_factorsVm(df, name):
    writer = pd.ExcelWriter('Test.xlsx', engine='openpyxl', mode='a')
    EBP_UPTIME = round(mean(df['UPTIMEPERCENTAGE'] * 100), 2)
    GAIN_ACHIEVEMENT = round(df['GAINACHIEVEMENT'] * 100, 2)
    GAIN_ACHIEVEMENT = GAIN_ACHIEVEMENT[1]
    REVENUE_ACHIEVEMENT = round(df.iloc[0]['REVENUEACHIEVEMENT'], 2)
    missingvalue(df)
    Oncall = sum(df['ONCALLS'])
    a = ['EXPECTGAIN', 'GAIN', 'GAINACHIEVEMENT',
     'REVENUEACHIEVEMENT', 'ANYCRITICALISSUE', 'DOWNTIMEINMINS',
     'REVENUEIMPACT', 'SCHEDULEDTIME', 'UPTIMEPERCENTAGE',
     'FAILEDROUTEIMPACT', 'ONAGENTSLAIMPACT', 'ONCALLSLAIMPACT',
     '1-1CALLSIMPACT', 'ONEVALUATIONERRIMPACT', 'LOOKUPFAILUREIMPACT',
     'UNKNOWNAGENTIMPACT', 'CGNOTFOUNDIMPACT', 'DISTINCTAGENTS',
     'PREVIOUSTOTALCALLS', 'CALLDIFFPERC', 'TOTALCALLS', 'ONCALLS',
     'OFFCALLS', 'BENCHMARK', 'SUCCESSROUTES', 'FAILROUTEPERC',
     'OFFAGENTSLA', 'OFFAGENTSLA%AGE', 'ONAGENTSLA', 'ONAGENTSLA%AGE',
     'OFFCALLSLA', 'OFFCALLSLA%AGE', 'ONCALLSLA', 'ONCALLSLA%AGE',
     '1-1CALLS', '1-1CALLS%AGE', '1-1CALLSWITHOUTSLABLOWNS',
     '1-1CALLS%AGEWITHOUTSLABLOWNS', 'L2CALLS', 'L2CALLS%AGE', 'O0BANDONS',
     'OFFABANDONS', 'O0BANDONSPERC', 'OFFABANDONSPERC', 'O0BAN-OFFABANPERC',
     'O0P', 'OFFAP', 'APSKEW', 'ONCP', 'OFFCP', 'AGENTCHOICE',
     'USEDAGENTCHOICE', 'USEDAGENTCHOICEWITHOUTSLABLOWNS', 'CALLCHOICE',
     'USEDCALLCHOICE', 'USEDCALLCHOICEWITHOUTSLABLOWNS', 'ONEVALSCORERAW',
     'OFFEVALSCORERAW', 'ONEVALSCOREUSED', 'OFFEVALSCOREUSED',
     'ONEVALUATIONERRCALLS', 'ONEVALUATIONERRCALLS%AGE',
     'OFFEVALUATIONERRCALLS', 'OFFEVALUATIONERRCALLS%AGE', 'LOOKUPFAILURES',
     'LOOKUPFAILUREPERC', 'UNKNOWNAGENTCALLS', 'UNKNOWNAGENTCALLS%AGE',
     'CGNOTFOUNDCALLS', 'CGNOTFOUNDCALLS%AGE',
     'IMPORTANT/HIGH/CRITICALTICKETS']
    Actual_Values = {}
    for i in a:
        Actual_Values[i] = round(sum(df[i]) * 100 / Oncall, 2)
    else:
        One_One_Calls_Actual = Actual_Values['1-1CALLS']
        One_One_Calls_Impact = round(df.iloc[0]['1-1CALLSIMPACT'] * 100, 2)
        Unknown_Agent_Calls_Actual = Actual_Values['UNKNOWNAGENTCALLS']
        Unknown_Agent_Calls_Impact = round(df.iloc[0]['UNKNOWNAGENTIMPACT'] * 100, 2)
        data1 = pd.DataFrame(Actual_Values.items())
        data2 = pd.DataFrame([['EBP_UPTIME', 'GAIN_ACHIEVEMENT', 'REVENUE_ACHIEVEMENT', 'One_One_Calls_Actual', 'One_One_Calls_Impact', 'Unknown_Agent_Calls_Actual', 'Unknown_Agent_Calls_Impact'], [EBP_UPTIME, GAIN_ACHIEVEMENT, REVENUE_ACHIEVEMENT, One_One_Calls_Actual, One_One_Calls_Impact, Unknown_Agent_Calls_Actual, Unknown_Agent_Calls_Impact]])
        data1 = data1.append(data2)
        if name == 'VMCare':
            data1.to_excel(writer, sheet_name='VMCare_Calculations')
            writer.save()
            writer.close()
        elif name == 'VM Cable Care':
            data1.to_excel(writer, sheet_name='VmCableCare_Calculations')
            writer.save()
            writer.close()


def get_factorsLM(df, name):
    writer = pd.ExcelWriter('Test.xlsx', engine='openpyxl', mode='a')
    EBP_UPTIME = round(mean(df['UPTIMEPERCENTAGE'] * 100), 2)
    GAIN_ACHIEVEMENT = round(df['GAINACHIEVEMENT'] * 100, 2)
    GAIN_ACHIEVEMENT = GAIN_ACHIEVEMENT[1]
    REVENUE_ACHIEVEMENT = round(df.iloc[0]['REVENUEACHIEVEMENT'], 2)
    missingvalue(df)
    Oncall = sum(df['ONCALLS'])
    a = ['ANYCRITICALISSUE', 'EXPECTGAIN', 'GAIN',
     'GAINACHIEVEMENT', 'REVENUEACHIEVEMENT', 'DOWNTIMEINMINS',
     'REVENUEIMPACT', 'DISTINCTAGENTS', 'SCHEDULEDTIME', 'UPTIMEPERCENTAGE',
     'FAILEDROUTEIMPACT', 'ONAGENTSLAIMPACT', 'ONCALLSLAIMPACT',
     '1-1CALLSIMPACT', 'ONEVALUATIONERRIMPACT', 'LOOKUPFAILUREIMPACT',
     'UNKNOWNAGENTIMPACT', 'CGNOTFOUNDIMPACT', 'PREVIOUSTOTALCALLS',
     'CALLDIFF%', 'TOTALCALLS', 'ONCALLS', 'OFFCALLS', 'ONBENCHMARK',
     'SUCCESSROUTES', 'FAILROUTEPERC', 'OFFAGENTSLA', 'OFFAGENTSLA%AGE',
     'ONAGENTSLA', 'ONAGENTSLA%AGE', 'OFFCALLSLA', 'OFFCALLSLA%AGE',
     'ONCALLSLA', 'ONCALLSLA%AGE', '1-1CALLS', '1-1CALLS%AGE',
     '1-1CALLSWITHOUTSLABLOWNS', '1-1CALLS%AGEWITHOUTSLABLOWNS', 'L2CALLS',
     'L2CALLS%AGE', 'O0BANDONS', 'OFFABANDONS', 'O0BANDONSPERC',
     'OFFABANDONSPERC', 'ON/OFFABANDONDIFF', 'O0P', 'OFFAP', 'APSKEW',
     'ONCP', 'OFFCP', 'AGENTCHOICE', 'USEDAGENTCHOICE',
     'USEDAGENTCHOICEWITHOUTSLABLOWNS', 'CALLCHOICE', 'USEDCALLCHOICE',
     'USEDCALLCHOICEWITHOUTSLABLOWNS', 'ONEVALSCORERAW', 'OFFEVALSCORERAW',
     'ONEVALSCOREUSED', 'OFFEVALSCOREUSED', 'ONEVALUATIONERRCALLS',
     'ONEVALUATIONERRCALLS%AGE', 'OFFEVALUATIONERRCALLS',
     'OFFEVALUATIONERRCALLS%AGE', 'LOOKUPFAILURES', 'LOOKUPFAILUREPERC',
     'UNKNOWNAGENTCALLS', 'UNKNOWNAGENTCALLS%AGE', 'CGNOTFOUNDCALLS',
     'CGNOTFOUNDCALLS%AGE', 'H/CISSUES']
    Actual_Values = {}
    for i in a:
        Actual_Values[i] = round(sum(df[i]) * 100 / Oncall, 2)
    else:
        One_One_Calls_Actual = Actual_Values['1-1CALLS']
        One_One_Calls_Impact = round(df.iloc[0]['1-1CALLSIMPACT'] * 100, 2)
        Unknown_Agent_Calls_Actual = Actual_Values['UNKNOWNAGENTCALLS']
        Unknown_Agent_Calls_Impact = round(df.iloc[0]['UNKNOWNAGENTIMPACT'] * 100, 2)
        data1 = pd.DataFrame(Actual_Values.items())
        data2 = pd.DataFrame([['EBP_UPTIME', 'GAIN_ACHIEVEMENT', 'REVENUE_ACHIEVEMENT', 'One_One_Calls_Actual', 'One_One_Calls_Impact', 'Unknown_Agent_Calls_Actual', 'Unknown_Agent_Calls_Impact'], [EBP_UPTIME, GAIN_ACHIEVEMENT, REVENUE_ACHIEVEMENT, One_One_Calls_Actual, One_One_Calls_Impact, Unknown_Agent_Calls_Actual, Unknown_Agent_Calls_Impact]])
        data1 = data1.append(data2)
        data1.to_excel(writer, sheet_name='LM_Calculations')
        writer.save()
        writer.close()


def get_factorsHouseCalls(df, name):
    writer = pd.ExcelWriter('Test.xlsx', engine='openpyxl', mode='a')
    EBP_UPTIME = round(mean(df['UPTIMEPERCENTAGE'] * 100), 2)
    GAIN_ACHIEVEMENT = round(df['GAINACHIEVEMENT'] * 100, 2)
    GAIN_ACHIEVEMENT = GAIN_ACHIEVEMENT[1]
    REVENUE_ACHIEVEMENT = round(df.iloc[0]['REVENUEACHIEVEMENT'], 2)
    missingvalue(df)
    Oncall = sum(df['ONCALLS'])
    a = [
     'ANYCRITICALISSUE', 'EXPECTGAIN', 'GAIN',
     'GAINACHIEVEMENT', 'REVENUEACHIEVEMENT', 'DOWNTIMEINMINS',
     'REVENUEIMPACT', 'DISTINCTAGENTS', 'SCHEDULEDTIME', 'UPTIMEPERCENTAGE',
     'FAILEDROUTEIMPACT', 'ONAGENTSLAIMPACT', 'ONCALLSLAIMPACT',
     '1-1CALLSIMPACT', 'ONEVALUATIONERRIMPACT', 'LOOKUPFAILUREIMPACT',
     'UNKNOWNAGENTIMPACT', 'CGNOTFOUNDIMPACT', 'PREVIOUSTOTALCALLS',
     'CALLDIFF%', 'TOTALCALLS', 'ONCALLS', 'OFFCALLS', 'ONBENCHMARK',
     'SUCCESSROUTES', 'FAILROUTEPERC', 'OFFAGENTSLA', 'OFFAGENTSLA%AGE',
     'ONAGENTSLA', 'ONAGENTSLA%AGE', 'OFFCALLSLA', 'OFFCALLSLA%AGE',
     'ONCALLSLA', 'ONCALLSLA%AGE', '1-1CALLS', '1-1CALLS%AGE',
     '1-1CALLSWITHOUTSLABLOWNS', '1-1CALLS%AGEWITHOUTSLABLOWNS', 'L2CALLS',
     'L2CALLS%AGE', 'O0BANDONS', 'OFFABANDONS', 'O0BANDONSPERC',
     'OFFABANDONSPERC', 'ON/OFFABANDONDIFF', 'O0P', 'OFFAP', 'APSKEW',
     'ONCP', 'OFFCP', 'AGENTCHOICE', 'USEDAGENTCHOICE',
     'USEDAGENTCHOICEWITHOUTSLABLOWNS', 'CALLCHOICE', 'USEDCALLCHOICE',
     'USEDCALLCHOICEWITHOUTSLABLOWNS', 'ONEVALSCORERAW', 'OFFEVALSCORERAW',
     'ONEVALSCOREUSED', 'OFFEVALSCOREUSED', 'ONEVALUATIONERRCALLS',
     'ONEVALUATIONERRCALLS%AGE', 'OFFEVALUATIONERRCALLS',
     'OFFEVALUATIONERRCALLS%AGE', 'LOOKUPFAILURES', 'LOOKUPFAILUREPERC',
     'UNKNOWNAGENTCALLS', 'UNKNOWNAGENTCALLS%AGE', 'CGNOTFOUNDCALLS',
     'CGNOTFOUNDCALLS%AGE', 'H/CISSUES']
    Actual_Values = {}
    for i in a:
        Actual_Values[i] = round(sum(df[i]) * 100 / Oncall, 2)
    else:
        One_One_Calls_Actual = Actual_Values['1-1CALLS']
        One_One_Calls_Impact = round(df.iloc[0]['1-1CALLSIMPACT'] * 100, 2)
        Unknown_Agent_Calls_Actual = Actual_Values['UNKNOWNAGENTCALLS']
        Unknown_Agent_Calls_Impact = round(df.iloc[0]['UNKNOWNAGENTIMPACT'] * 100, 2)
        data1 = pd.DataFrame(Actual_Values.items())
        data2 = pd.DataFrame([['EBP_UPTIME', 'GAIN_ACHIEVEMENT', 'REVENUE_ACHIEVEMENT', 'One_One_Calls_Actual', 'One_One_Calls_Impact', 'Unknown_Agent_Calls_Actual', 'Unknown_Agent_Calls_Impact'], [EBP_UPTIME, GAIN_ACHIEVEMENT, REVENUE_ACHIEVEMENT, One_One_Calls_Actual, One_One_Calls_Impact, Unknown_Agent_Calls_Actual, Unknown_Agent_Calls_Impact]])
        data1 = data1.append(data2)
        data1.to_excel(writer, sheet_name='HouseCalls_Calculations')
        writer.save()
        writer.close()


def get_factorsNurseLine(df, name):
    writer = pd.ExcelWriter('Test.xlsx', engine='openpyxl', mode='a')
    EBP_UPTIME = round(mean(df['UPTIMEPERCENTAGE'] * 100), 2)
    GAIN_ACHIEVEMENT = round(df['GAINACHIEVEMENT'] * 100, 2)
    GAIN_ACHIEVEMENT = GAIN_ACHIEVEMENT[1]
    REVENUE_ACHIEVEMENT = round(df.iloc[0]['REVENUEACHIEVEMENT'], 2)
    missingvalue(df)
    Oncall = sum(df['ONCALLS'])
    a = [
     'ANYCRITICALISSUE', 'EXPECTGAIN', 'GAIN',
     'GAINACHIEVEMENT', 'REVENUEACHIEVEMENT', 'DOWNTIMEINMINS',
     'REVENUEIMPACT', 'DISTINCTAGENTS', 'SCHEDULEDTIME', 'UPTIMEPERCENTAGE',
     'FAILEDROUTEIMPACT', 'ONAGENTSLAIMPACT', 'ONCALLSLAIMPACT',
     '1-1CALLSIMPACT', 'ONEVALUATIONERRIMPACT', 'LOOKUPFAILUREIMPACT',
     'UNKNOWNAGENTIMPACT', 'CGNOTFOUNDIMPACT', 'PREVIOUSTOTALCALLS',
     'CALLDIFF%', 'TOTALCALLS', 'ONCALLS', 'OFFCALLS', 'ONBENCHMARK',
     'SUCCESSROUTES', 'FAILROUTEPERC', 'OFFAGENTSLA', 'OFFAGENTSLA%AGE',
     'ONAGENTSLA', 'ONAGENTSLA%AGE', 'OFFCALLSLA', 'OFFCALLSLA%AGE',
     'ONCALLSLA', 'ONCALLSLA%AGE', '1-1CALLS', '1-1CALLS%AGE',
     '1-1CALLSWITHOUTSLABLOWNS', '1-1CALLS%AGEWITHOUTSLABLOWNS', 'L2CALLS',
     'L2CALLS%AGE', 'O0BANDONS', 'OFFABANDONS', 'O0BANDONSPERC',
     'OFFABANDONSPERC', 'ON/OFFABANDONDIFF', 'O0P', 'OFFAP', 'APSKEW',
     'ONCP', 'OFFCP', 'AGENTCHOICE', 'USEDAGENTCHOICE',
     'USEDAGENTCHOICEWITHOUTSLABLOWNS', 'CALLCHOICE', 'USEDCALLCHOICE',
     'USEDCALLCHOICEWITHOUTSLABLOWNS', 'ONEVALSCORERAW', 'OFFEVALSCORERAW',
     'ONEVALSCOREUSED', 'OFFEVALSCOREUSED', 'ONEVALUATIONERRCALLS',
     'ONEVALUATIONERRCALLS%AGE', 'OFFEVALUATIONERRCALLS',
     'OFFEVALUATIONERRCALLS%AGE', 'LOOKUPFAILURES', 'LOOKUPFAILUREPERC',
     'UNKNOWNAGENTCALLS', 'UNKNOWNAGENTCALLS%AGE', 'CGNOTFOUNDCALLS',
     'CGNOTFOUNDCALLS%AGE']
    Actual_Values = {}
    for i in a:
        Actual_Values[i] = round(sum(df[i]) * 100 / Oncall, 2)
    else:
        One_One_Calls_Actual = Actual_Values['1-1CALLS']
        One_One_Calls_Impact = round(df.iloc[0]['1-1CALLSIMPACT'] * 100, 2)
        Unknown_Agent_Calls_Actual = Actual_Values['UNKNOWNAGENTCALLS']
        Unknown_Agent_Calls_Impact = round(df.iloc[0]['UNKNOWNAGENTIMPACT'] * 100, 2)
        data1 = pd.DataFrame(Actual_Values.items())
        data2 = pd.DataFrame([['EBP_UPTIME', 'GAIN_ACHIEVEMENT', 'REVENUE_ACHIEVEMENT', 'One_One_Calls_Actual', 'One_One_Calls_Impact', 'Unknown_Agent_Calls_Actual', 'Unknown_Agent_Calls_Impact'], [EBP_UPTIME, GAIN_ACHIEVEMENT, REVENUE_ACHIEVEMENT, One_One_Calls_Actual, One_One_Calls_Impact, Unknown_Agent_Calls_Actual, Unknown_Agent_Calls_Impact]])
        data1 = data1.append(data2)
        data1.to_excel(writer, sheet_name='NurseLine_Calculations')
        writer.save()
        writer.close()


def get_factorsATT(df, name):
    writer = pd.ExcelWriter('Test.xlsx', engine='openpyxl', mode='a')
    EBP_UPTIME = round(mean(df['UPTIMEPERCENTAGE'] * 100), 2)
    GAIN_ACHIEVEMENT = round(df['GAINACHIEVEMENT'] * 100, 2)
    GAIN_ACHIEVEMENT = GAIN_ACHIEVEMENT[1]
    REVENUE_ACHIEVEMENT = round(df.iloc[0]['REVENUEACHIEVEMENT'], 2)
    missingvalue(df)
    Oncall = sum(df['ONCALLS'])
    a = [
     'PROGRAM', 'DATE', 'ANYCRITICALISSUE', 'EXPECTGAIN', 'GAIN',
     'GAINACHIEVEMENT', 'REVENUEACHIEVEMENT', 'DOWNTIMEINMINS',
     'REVENUEIMPACT', 'DISTINCTAGENTS', 'SCHEDULEDTIME', 'UPTIMEPERCENTAGE',
     'FAILEDROUTEIMPACT', 'ONAGENTSLAIMPACT', 'ONCALLSLAIMPACT',
     '1-1CALLSIMPACT', 'ONEVALUATIONERRIMPACT', 'LOOKUPFAILUREIMPACT',
     'UNKNOWNAGENTIMPACT', 'CGNOTFOUNDIMPACT', 'PREVIOUSTOTALCALLS',
     'CALLDIFFPERC', 'TOTALCALLS', 'ONCALLS', 'OFFCALLS', 'BENCHMARK',
     'SUCCESSROUTES', 'FAILROUTEPERC', 'OFFAGENTSLA', 'OFFAGENTSLA%AGE',
     'ONAGENTSLA', 'ONAGENTSLA%AGE', 'OFFCALLSLA', 'OFFCALLSLA%AGE',
     'ONCALLSLA', 'ONCALLSLA%AGE', '1-1CALLS', '1-1CALLS%AGE',
     '1-1CALLSWITHOUTSLABLOWNS', '1-1CALLS%AGEWITHOUTSLABLOWNS', 'L2CALLS',
     'L2CALLS%AGE', 'O0BANDONS', 'OFFABANDONS', 'O0BANDONSPERC',
     'OFFABANDONSPERC', 'O0BAN-OFFABANPERC', 'O0P', 'OFFAP', 'APSKEW',
     'ONCP', 'OFFCP', 'AGENTCHOICE', 'USEDAGENTCHOICE',
     'USEDAGENTCHOICEWITHOUTSLABLOWNS', 'CALLCHOICE', 'USEDCALLCHOICE',
     'USEDCALLCHOICEWITHOUTSLABLOWNS', 'ONEVALSCORERAW', 'OFFEVALSCORERAW',
     'ONEVALSCOREUSED', 'OFFEVALSCOREUSED', 'ONEVALUATIONERRCALLS',
     'ONEVALUATIONERRCALLS%AGE', 'OFFEVALUATIONERRCALLS',
     'OFFEVALUATIONERRCALLS%AGE', 'LOOKUPFAILURES', 'LOOKUPFAILUREPERC',
     'UNKNOWNAGENTCALLS', 'UNKNOWNAGENTCALLS%AGE', 'CGNOTFOUNDCALLS',
     'CGNOTFOUNDCALLS%AGE']
    Actual_Values = {}
    for i in a:
        Actual_Values[i] = round(sum(df[i]) * 100 / Oncall, 2)
    else:
        One_One_Calls_Actual = Actual_Values['1-1CALLS']
        One_One_Calls_Impact = round(df.iloc[0]['1-1CALLSIMPACT'] * 100, 2)
        Unknown_Agent_Calls_Actual = Actual_Values['UNKNOWNAGENTCALLS']
        Unknown_Agent_Calls_Impact = round(df.iloc[0]['UNKNOWNAGENTIMPACT'] * 100, 2)
        data1 = pd.DataFrame(Actual_Values.items())
        data2 = pd.DataFrame([['EBP_UPTIME', 'GAIN_ACHIEVEMENT', 'REVENUE_ACHIEVEMENT', 'One_One_Calls_Actual', 'One_One_Calls_Impact', 'Unknown_Agent_Calls_Actual', 'Unknown_Agent_Calls_Impact'], [EBP_UPTIME, GAIN_ACHIEVEMENT, REVENUE_ACHIEVEMENT, One_One_Calls_Actual, One_One_Calls_Impact, Unknown_Agent_Calls_Actual, Unknown_Agent_Calls_Impact]])
        data1 = data1.append(data2)
        if name == 'AT&TPhase3_GenDTVSales':
            data1.to_excel(writer, sheet_name='AT&TPhase3_GenDTVSales_Calculations')
            writer.save()
            writer.close()
        elif name == 'AT&TPhase3_GenHSCLG':
            data1.to_excel(writer, sheet_name='AT&TPhase3_GenHSCLG_Calculations')
            writer.save()
            writer.close()
        elif name == 'AT&TPhase3_GENHSSALES':
            data1.to_excel(writer, sheet_name='AT&TPhase3_GENHSSALES_Calculations')
            writer.save()
            writer.close()
        elif name == 'AT&TPhase4_DMDR':
            data1.to_excel(writer, sheet_name='AT&TPhase4_DMDR_Calculations')
            writer.save()
            writer.close()
        elif name == 'AT&TPhase4_ISMCLG':
            data1.to_excel(writer, sheet_name='AT&TPhase4_ISMCLG_Calculations')
            writer.save()
            writer.close()
        elif name == 'AT&TPhase4_ISMSVC':
            data1.to_excel(writer, sheet_name='AT&TPhase4_ISMSVC_Calculations')
            writer.save()
            writer.close()
        elif name == 'AT&TPhase4_MobCLG':
            data1.to_excel(writer, sheet_name='AT&TPhase4_MobCLG_Calculations')
            writer.save()
            writer.close()
        elif name == 'AT&TPhase4_MobSS':
            data1.to_excel(writer, sheet_name='AT&TPhase4_MobSS_Calculations')
            writer.save()
            writer.close()


def missingvalue(df):
    a = df.columns.tolist()
    for i in a:
        df[i] = pd.to_numeric((df[i]), errors='coerce').fillna(0)
    else:
        return df


def write_excel(filename, sheetname, dataframe):
    with pd.ExcelWriter(filename, engine='openpyxl', mode='a') as writer:
        workBook = writer.book
        try:
            try:
                workBook.remove(workBook[sheetname])
            except:
                print('Worksheet does not exist')

        finally:
            dataframe.to_excel(writer, sheet_name=sheetname, index=False)
            writer.save()


def insert_record(text, e, g, sched):
    global LM
    global Verizon
    global care
    global housecalls
    global nurseline
    global phase3_gendhsclg
    global phase3_gendhssales
    global phase3_gendtvsales
    global phase4_dmdr
    global phase4_ismclg
    global phase4_ismsvc
    global phase4_mobclg
    global phase4_mobss
    global sales
    global sky
    global skyservice
    global vmcablecare
    global vmcare
    e = float(e)
    g = float(g)
    sched = int(sched)
    if text == 'Bygs_Sales':
        sales = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='Bygs_Sales')
        sales.columns = map(lambda x: str(x).upper(), sales.columns)
        sales.columns = map(lambda x: str(x).replace('_', ''), sales.columns)
        sales.columns = map(lambda x: str(x).replace(' ', ''), sales.columns)
        sales.columns = map(lambda x: str(x).rstrip(), sales.columns)
        sales.insert(3, 'EXPECTGAIN', e)
        sales.insert(4, 'GAIN', g)
        summ = sales.sum(axis=0)
        sales.insert(5, 'GAINACHIEVEMENT', round(sales['GAIN'] / sales['EXPECTGAIN'] * 100, 2))
        sales.insert(9, 'SCHEDULEDTIME', sched)
        sc = sales['SCHEDULEDTIME']
        sales['DOWNTIMEINMINS'] = pd.to_numeric((sales['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = sales['DOWNTIMEINMINS']
        a = sc - dm
        sales.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(sales['UPTIMEPERCENTAGE'] * 100), 2)
        sales.insert(6, 'REVENUEACHIEVEMENT', round(sales['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        sales.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        sales.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        sales.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        sales.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        sales.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        sales.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        sales.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        sales.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, sales)
        get_factors(sales, 'Bygs_Sales')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')
    elif text == 'Bygs_Care':
        care = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='Bygs_Care')
        care.columns = map(lambda x: str(x).upper(), care.columns)
        care.columns = map(lambda x: str(x).replace('_', ''), care.columns)
        care.columns = map(lambda x: str(x).replace(' ', ''), care.columns)
        care.columns = map(lambda x: str(x).rstrip(), care.columns)
        care.insert(3, 'EXPECTGAIN', e)
        care.insert(4, 'GAIN', g)
        summ = care.sum(axis=0)
        care.insert(5, 'GAINACHIEVEMENT', round(care['GAIN'] / care['EXPECTGAIN'] * 100, 2))
        care.insert(9, 'SCHEDULEDTIME', sched)
        sc = care['SCHEDULEDTIME']
        care['DOWNTIMEINMINS'] = pd.to_numeric((care['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = care['DOWNTIMEINMINS']
        a = sc - dm
        care.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(care['UPTIMEPERCENTAGE'] * 100), 2)
        care.insert(6, 'REVENUEACHIEVEMENT', round(care['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        care.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        care.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        care.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        care.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        care.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        care.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        care.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        care.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, care)
        get_factors(care, 'Bygs_Care')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')
    elif text == 'Verizon':
        Verizon = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='Verizon')
        Verizon.columns = map(lambda x: str(x).upper(), Verizon.columns)
        Verizon.columns = map(lambda x: str(x).replace('_', ''), Verizon.columns)
        Verizon.columns = map(lambda x: str(x).replace(' ', ''), Verizon.columns)
        Verizon.columns = map(lambda x: str(x).rstrip(), Verizon.columns)
        Verizon.insert(3, 'EXPECTGAIN', e)
        Verizon.insert(4, 'GAIN', g)
        summ = Verizon.sum(axis=0)
        Verizon.insert(5, 'GAINACHIEVEMENT', round(Verizon['GAIN'] / Verizon['EXPECTGAIN'] * 100, 2))
        Verizon.insert(9, 'SCHEDULEDTIME', sched)
        sc = Verizon['SCHEDULEDTIME']
        Verizon['DOWNTIMEINMINS'] = pd.to_numeric((Verizon['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = Verizon['DOWNTIMEINMINS']
        a = sc - dm
        Verizon.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(Verizon['UPTIMEPERCENTAGE'] * 100), 2)
        Verizon.insert(6, 'REVENUEACHIEVEMENT', round(Verizon['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        Verizon.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        Verizon.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        Verizon.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        Verizon.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        Verizon.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        Verizon.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        Verizon.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        Verizon.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, Verizon)
        get_factors(Verizon, 'Verizon')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')
    elif text == 'SkyService':
        skyservice = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='SkyService')
        skyservice.columns = map(lambda x: str(x).upper(), skyservice.columns)
        skyservice.columns = map(lambda x: str(x).replace('_', ''), skyservice.columns)
        skyservice.columns = map(lambda x: str(x).replace(' ', ''), skyservice.columns)
        skyservice.columns = map(lambda x: str(x).rstrip(), skyservice.columns)
        skyservice.insert(3, 'EXPECTGAIN', e)
        skyservice.insert(4, 'GAIN', g)
        summ = skyservice.sum(axis=0)
        skyservice.insert(5, 'GAINACHIEVEMENT', round(skyservice['GAIN'] / skyservice['EXPECTGAIN'] * 100, 2))
        skyservice.insert(9, 'SCHEDULEDTIME', sched)
        sc = skyservice['SCHEDULEDTIME']
        skyservice['DOWNTIMEINMINS'] = pd.to_numeric((skyservice['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = skyservice['DOWNTIMEINMINS']
        a = sc - dm
        skyservice.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(skyservice['UPTIMEPERCENTAGE'] * 100), 2)
        skyservice.insert(6, 'REVENUEACHIEVEMENT', round(skyservice['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        skyservice.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        skyservice.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        skyservice.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        skyservice.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        skyservice.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        skyservice.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        skyservice.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        skyservice.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, skyservice)
        get_factorsSkyService(skyservice, 'SkyService')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')
    elif text == 'SKY_PlatformClassic':
        sky = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='SKY_PlatformClassic')
        sky.columns = map(lambda x: str(x).upper(), sky.columns)
        sky.columns = map(lambda x: str(x).replace('_', ''), sky.columns)
        sky.columns = map(lambda x: str(x).replace(' ', ''), sky.columns)
        sky.columns = map(lambda x: str(x).rstrip(), sky.columns)
        sky.insert(3, 'EXPECTGAIN', e)
        sky.insert(4, 'GAIN', g)
        summ = sky.sum(axis=0)
        sky.insert(5, 'GAINACHIEVEMENT', round(sky['GAIN'] / sky['EXPECTGAIN'] * 100, 2))
        sky.insert(9, 'SCHEDULEDTIME', sched)
        sc = sky['SCHEDULEDTIME']
        sky['DOWNTIMEINMINS'] = pd.to_numeric((sky['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = sky['DOWNTIMEINMINS']
        a = sc - dm
        sky.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(sky['UPTIMEPERCENTAGE'] * 100), 2)
        sky.insert(6, 'REVENUEACHIEVEMENT', round(sky['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        sky.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        sky.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        sky.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        sky.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        sky.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        sky.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        sky.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        sky.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, sky)
        get_factorsSky(sky, 'SKY_PlatformClassic')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')
    elif text == 'LM':
        LM = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='LM')
        LM.columns = map(lambda x: str(x).upper(), LM.columns)
        LM.columns = map(lambda x: str(x).replace('_', ''), LM.columns)
        LM.columns = map(lambda x: str(x).replace(' ', ''), LM.columns)
        LM.columns = map(lambda x: str(x).rstrip(), LM.columns)
        LM.insert(3, 'EXPECTGAIN', e)
        LM.insert(4, 'GAIN', g)
        summ = LM.sum(axis=0)
        LM.insert(5, 'GAINACHIEVEMENT', round(LM['GAIN'] / LM['EXPECTGAIN'] * 100, 2))
        LM.insert(9, 'SCHEDULEDTIME', sched)
        sc = LM['SCHEDULEDTIME']
        LM['DOWNTIMEINMINS'] = pd.to_numeric((LM['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = LM['DOWNTIMEINMINS']
        a = sc - dm
        LM.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(LM['UPTIMEPERCENTAGE'] * 100), 2)
        LM.insert(6, 'REVENUEACHIEVEMENT', round(LM['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        LM.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        LM.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        LM.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        LM.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        LM.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        LM.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        LM.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        LM.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, LM)
        get_factorsLM(LM, 'LM')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')
    elif text == 'VMCare':
        vmcare = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='VMCare')
        vmcare.columns = map(lambda x: str(x).upper(), vmcare.columns)
        vmcare.columns = map(lambda x: str(x).replace('_', ''), vmcare.columns)
        vmcare.columns = map(lambda x: str(x).replace(' ', ''), vmcare.columns)
        vmcare.columns = map(lambda x: str(x).rstrip(), vmcare.columns)
        vmcare.insert(3, 'EXPECTGAIN', e)
        vmcare.insert(4, 'GAIN', g)
        summ = vmcare.sum(axis=0)
        vmcare.insert(5, 'GAINACHIEVEMENT', round(vmcare['GAIN'] / vmcare['EXPECTGAIN'] * 100, 2))
        vmcare.insert(9, 'SCHEDULEDTIME', sched)
        sc = vmcare['SCHEDULEDTIME']
        vmcare['DOWNTIMEINMINS'] = pd.to_numeric((vmcare['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = vmcare['DOWNTIMEINMINS']
        a = sc - dm
        vmcare.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(vmcare['UPTIMEPERCENTAGE'] * 100), 2)
        vmcare.insert(6, 'REVENUEACHIEVEMENT', round(vmcare['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        vmcare.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        vmcare.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        vmcare.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        vmcare.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        vmcare.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        vmcare.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        vmcare.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        vmcare.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, vmcare)
        get_factorsVm(vmcare, 'VMCare')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')
    elif text == 'VM Cable Care':
        vmcablecare = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='VM Cable Care')
        vmcablecare.columns = map(lambda x: str(x).upper(), vmcablecare.columns)
        vmcablecare.columns = map(lambda x: str(x).replace('_', ''), vmcablecare.columns)
        vmcablecare.columns = map(lambda x: str(x).replace(' ', ''), vmcablecare.columns)
        vmcablecare.columns = map(lambda x: str(x).rstrip(), vmcablecare.columns)
        vmcablecare.insert(3, 'EXPECTGAIN', e)
        vmcablecare.insert(4, 'GAIN', g)
        summ = vmcablecare.sum(axis=0)
        vmcablecare.insert(5, 'GAINACHIEVEMENT', round(vmcablecare['GAIN'] / vmcablecare['EXPECTGAIN'] * 100, 2))
        vmcablecare.insert(9, 'SCHEDULEDTIME', sched)
        sc = vmcablecare['SCHEDULEDTIME']
        vmcablecare['DOWNTIMEINMINS'] = pd.to_numeric((vmcablecare['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = vmcablecare['DOWNTIMEINMINS']
        a = sc - dm
        vmcablecare.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(vmcablecare['UPTIMEPERCENTAGE'] * 100), 2)
        vmcablecare.insert(6, 'REVENUEACHIEVEMENT', round(vmcablecare['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        vmcablecare.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        vmcablecare.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        vmcablecare.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        vmcablecare.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        vmcablecare.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        vmcablecare.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        vmcablecare.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        vmcablecare.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, vmcablecare)
        get_factorsVm(vmcablecare, 'VM Cable Care')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')
    elif text == 'AT&TPhase3_GenDTVSales':
        phase3_gendtvsales = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='AT&TPhase3_GenDTVSales')
        phase3_gendtvsales.columns = map(lambda x: str(x).upper(), phase3_gendtvsales.columns)
        phase3_gendtvsales.columns = map(lambda x: str(x).replace('_', ''), phase3_gendtvsales.columns)
        phase3_gendtvsales.columns = map(lambda x: str(x).replace(' ', ''), phase3_gendtvsales.columns)
        phase3_gendtvsales.columns = map(lambda x: str(x).rstrip(), phase3_gendtvsales.columns)
        phase3_gendtvsales.insert(3, 'EXPECTGAIN', e)
        phase3_gendtvsales.insert(4, 'GAIN', g)
        summ = phase3_gendtvsales.sum(axis=0)
        phase3_gendtvsales.insert(5, 'GAINACHIEVEMENT', round(phase3_gendtvsales['GAIN'] / phase3_gendtvsales['EXPECTGAIN'] * 100, 2))
        phase3_gendtvsales.insert(9, 'SCHEDULEDTIME', sched)
        sc = phase3_gendtvsales['SCHEDULEDTIME']
        phase3_gendtvsales['DOWNTIMEINMINS'] = pd.to_numeric((phase3_gendtvsales['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = phase3_gendtvsales['DOWNTIMEINMINS']
        a = sc - dm
        phase3_gendtvsales.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(phase3_gendtvsales['UPTIMEPERCENTAGE'] * 100), 2)
        phase3_gendtvsales.insert(6, 'REVENUEACHIEVEMENT', round(phase3_gendtvsales['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        phase3_gendtvsales.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendtvsales.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendtvsales.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendtvsales.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendtvsales.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendtvsales.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendtvsales.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendtvsales.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, phase3_gendtvsales)
        get_factorsATT(phase3_gendtvsales, 'AT&TPhase3_GenDTVSales')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')
    elif text == 'AT&TPhase3_GenHSCLG':
        phase3_gendhsclg = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='AT&TPhase3_GenHSCLG')
        phase3_gendhsclg.columns = map(lambda x: str(x).upper(), phase3_gendhsclg.columns)
        phase3_gendhsclg.columns = map(lambda x: str(x).replace('_', ''), phase3_gendhsclg.columns)
        phase3_gendhsclg.columns = map(lambda x: str(x).replace(' ', ''), phase3_gendhsclg.columns)
        phase3_gendhsclg.columns = map(lambda x: str(x).rstrip(), phase3_gendhsclg.columns)
        phase3_gendhsclg.insert(3, 'EXPECTGAIN', e)
        phase3_gendhsclg.insert(4, 'GAIN', g)
        summ = phase3_gendhsclg.sum(axis=0)
        phase3_gendhsclg.insert(5, 'GAINACHIEVEMENT', round(phase3_gendhsclg['GAIN'] / phase3_gendhsclg['EXPECTGAIN'] * 100, 2))
        phase3_gendhsclg.insert(9, 'SCHEDULEDTIME', sched)
        sc = phase3_gendhsclg['SCHEDULEDTIME']
        phase3_gendhsclg['DOWNTIMEINMINS'] = pd.to_numeric((phase3_gendhsclg['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = phase3_gendhsclg['DOWNTIMEINMINS']
        a = sc - dm
        phase3_gendhsclg.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(phase3_gendhsclg['UPTIMEPERCENTAGE'] * 100), 2)
        phase3_gendhsclg.insert(6, 'REVENUEACHIEVEMENT', round(phase3_gendhsclg['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        phase3_gendhsclg.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendhsclg.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendhsclg.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendhsclg.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendhsclg.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendhsclg.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendhsclg.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendhsclg.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, phase3_gendhsclg)
        get_factorsATT(phase3_gendhsclg, 'AT&TPhase3_GenHSCLG')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')
    elif text == 'AT&TPhase3_GENHSSALES':
        phase3_gendhssales = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='AT&TPhase3_GENHSSALES')
        phase3_gendhssales.columns = map(lambda x: str(x).upper(), phase3_gendhssales.columns)
        phase3_gendhssales.columns = map(lambda x: str(x).replace('_', ''), phase3_gendhssales.columns)
        phase3_gendhssales.columns = map(lambda x: str(x).replace(' ', ''), phase3_gendhssales.columns)
        phase3_gendhssales.columns = map(lambda x: str(x).rstrip(), phase3_gendhssales.columns)
        phase3_gendhssales.insert(3, 'EXPECTGAIN', e)
        phase3_gendhssales.insert(4, 'GAIN', g)
        summ = phase3_gendhssales.sum(axis=0)
        phase3_gendhssales.insert(5, 'GAINACHIEVEMENT', round(phase3_gendhssales['GAIN'] / phase3_gendhssales['EXPECTGAIN'] * 100, 2))
        phase3_gendhssales.insert(9, 'SCHEDULEDTIME', sched)
        sc = phase3_gendhssales['SCHEDULEDTIME']
        phase3_gendhssales['DOWNTIMEINMINS'] = pd.to_numeric((phase3_gendhssales['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = phase3_gendhssales['DOWNTIMEINMINS']
        a = sc - dm
        phase3_gendhssales.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(phase3_gendhssales['UPTIMEPERCENTAGE'] * 100), 2)
        phase3_gendhssales.insert(6, 'REVENUEACHIEVEMENT', round(phase3_gendhssales['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        phase3_gendhssales.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendhssales.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendhssales.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendhssales.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendhssales.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendhssales.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendhssales.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase3_gendhssales.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, phase3_gendhssales)
        get_factorsATT(phase3_gendhssales, 'AT&TPhase3_GENHSSALES')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')
    elif text == 'AT&TPhase4_DMDR':
        phase4_dmdr = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='AT&TPhase4_DMDR')
        phase4_dmdr.columns = map(lambda x: str(x).upper(), phase4_dmdr.columns)
        phase4_dmdr.columns = map(lambda x: str(x).replace('_', ''), phase4_dmdr.columns)
        phase4_dmdr.columns = map(lambda x: str(x).replace(' ', ''), phase4_dmdr.columns)
        phase4_dmdr.columns = map(lambda x: str(x).rstrip(), phase4_dmdr.columns)
        phase4_dmdr.insert(3, 'EXPECTGAIN', e)
        phase4_dmdr.insert(4, 'GAIN', g)
        summ = phase4_dmdr.sum(axis=0)
        phase4_dmdr.insert(5, 'GAINACHIEVEMENT', round(phase4_dmdr['GAIN'] / phase4_dmdr['EXPECTGAIN'] * 100, 2))
        phase4_dmdr.insert(9, 'SCHEDULEDTIME', sched)
        sc = phase4_dmdr['SCHEDULEDTIME']
        phase4_dmdr['DOWNTIMEINMINS'] = pd.to_numeric((phase4_dmdr['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = phase4_dmdr['DOWNTIMEINMINS']
        a = sc - dm
        phase4_dmdr.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(phase4_dmdr['UPTIMEPERCENTAGE'] * 100), 2)
        phase4_dmdr.insert(6, 'REVENUEACHIEVEMENT', round(phase4_dmdr['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        phase4_dmdr.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        phase4_dmdr.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_dmdr.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_dmdr.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        phase4_dmdr.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_dmdr.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        phase4_dmdr.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_dmdr.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, phase4_dmdr)
        get_factorsATT(phase4_dmdr, 'AT&TPhase4_DMDR')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')
    elif text == 'AT&TPhase4_ISMCLG':
        phase4_ismclg = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='AT&TPhase4_ISMCLG')
        phase4_ismclg.columns = map(lambda x: str(x).upper(), phase4_ismclg.columns)
        phase4_ismclg.columns = map(lambda x: str(x).replace('_', ''), phase4_ismclg.columns)
        phase4_ismclg.columns = map(lambda x: str(x).replace(' ', ''), phase4_ismclg.columns)
        phase4_ismclg.columns = map(lambda x: str(x).rstrip(), phase4_ismclg.columns)
        phase4_ismclg.insert(3, 'EXPECTGAIN', e)
        phase4_ismclg.insert(4, 'GAIN', g)
        summ = phase4_ismclg.sum(axis=0)
        phase4_ismclg.insert(5, 'GAINACHIEVEMENT', round(phase4_ismclg['GAIN'] / phase4_ismclg['EXPECTGAIN'] * 100, 2))
        phase4_ismclg.insert(9, 'SCHEDULEDTIME', sched)
        sc = phase4_ismclg['SCHEDULEDTIME']
        phase4_ismclg['DOWNTIMEINMINS'] = pd.to_numeric((phase4_ismclg['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = phase4_ismclg['DOWNTIMEINMINS']
        a = sc - dm
        phase4_ismclg.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(phase4_ismclg['UPTIMEPERCENTAGE'] * 100), 2)
        phase4_ismclg.insert(6, 'REVENUEACHIEVEMENT', round(phase4_ismclg['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        phase4_ismclg.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        phase4_ismclg.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_ismclg.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_ismclg.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        phase4_ismclg.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_ismclg.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        phase4_ismclg.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_ismclg.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, phase4_ismclg)
        get_factorsATT(phase4_ismclg, 'AT&TPhase4_ISMCLG')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')
    elif text == 'AT&TPhase4_ISMSVC':
        phase4_ismsvc = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='AT&TPhase4_ISMSVC')
        phase4_ismsvc.columns = map(lambda x: str(x).upper(), phase4_ismsvc.columns)
        phase4_ismsvc.columns = map(lambda x: str(x).replace('_', ''), phase4_ismsvc.columns)
        phase4_ismsvc.columns = map(lambda x: str(x).replace(' ', ''), phase4_ismsvc.columns)
        phase4_ismsvc.columns = map(lambda x: str(x).rstrip(), phase4_ismsvc.columns)
        phase4_ismsvc.insert(3, 'EXPECTGAIN', e)
        phase4_ismsvc.insert(4, 'GAIN', g)
        summ = phase4_ismsvc.sum(axis=0)
        phase4_ismsvc.insert(5, 'GAINACHIEVEMENT', round(phase4_ismsvc['GAIN'] / phase4_ismsvc['EXPECTGAIN'] * 100, 2))
        phase4_ismsvc.insert(9, 'SCHEDULEDTIME', sched)
        sc = phase4_ismsvc['SCHEDULEDTIME']
        phase4_ismsvc['DOWNTIMEINMINS'] = pd.to_numeric((phase4_ismsvc['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = phase4_ismsvc['DOWNTIMEINMINS']
        a = sc - dm
        phase4_ismsvc.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(phase4_ismsvc['UPTIMEPERCENTAGE'] * 100), 2)
        phase4_ismsvc.insert(6, 'REVENUEACHIEVEMENT', round(phase4_ismsvc['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        phase4_ismsvc.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        phase4_ismsvc.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_ismsvc.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_ismsvc.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        phase4_ismsvc.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_ismsvc.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        phase4_ismsvc.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_ismsvc.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, phase4_ismsvc)
        get_factorsATT(phase4_ismsvc, 'AT&TPhase4_ISMSVC')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')
    elif text == 'AT&TPhase4_MobCLG':
        phase4_mobclg = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='AT&TPhase4_MobCLG')
        phase4_mobclg.columns = map(lambda x: str(x).upper(), phase4_mobclg.columns)
        phase4_mobclg.columns = map(lambda x: str(x).replace('_', ''), phase4_mobclg.columns)
        phase4_mobclg.columns = map(lambda x: str(x).replace(' ', ''), phase4_mobclg.columns)
        phase4_mobclg.columns = map(lambda x: str(x).rstrip(), phase4_mobclg.columns)
        phase4_mobclg.insert(3, 'EXPECTGAIN', e)
        phase4_mobclg.insert(4, 'GAIN', g)
        summ = phase4_mobclg.sum(axis=0)
        phase4_mobclg.insert(5, 'GAINACHIEVEMENT', round(phase4_mobclg['GAIN'] / phase4_mobclg['EXPECTGAIN'] * 100, 2))
        phase4_mobclg.insert(9, 'SCHEDULEDTIME', sched)
        sc = phase4_mobclg['SCHEDULEDTIME']
        phase4_mobclg['DOWNTIMEINMINS'] = pd.to_numeric((phase4_mobclg['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = phase4_mobclg['DOWNTIMEINMINS']
        a = sc - dm
        phase4_mobclg.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(phase4_mobclg['UPTIMEPERCENTAGE'] * 100), 2)
        phase4_mobclg.insert(6, 'REVENUEACHIEVEMENT', round(phase4_mobclg['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        phase4_mobclg.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        phase4_mobclg.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_mobclg.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_mobclg.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        phase4_mobclg.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_mobclg.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        phase4_mobclg.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_mobclg.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, phase4_mobclg)
        get_factorsATT(phase4_mobclg, 'AT&TPhase4_MobCLG')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')
    elif text == 'AT&TPhase4_MobSS':
        phase4_mobss = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='AT&TPhase4_MobSS')
        phase4_mobss.columns = map(lambda x: str(x).upper(), phase4_mobss.columns)
        phase4_mobss.columns = map(lambda x: str(x).replace('_', ''), phase4_mobss.columns)
        phase4_mobss.columns = map(lambda x: str(x).replace(' ', ''), phase4_mobss.columns)
        phase4_mobss.columns = map(lambda x: str(x).rstrip(), phase4_mobss.columns)
        phase4_mobss.insert(3, 'EXPECTGAIN', e)
        phase4_mobss.insert(4, 'GAIN', g)
        summ = phase4_mobss.sum(axis=0)
        phase4_mobss.insert(5, 'GAINACHIEVEMENT', round(phase4_mobss['GAIN'] / phase4_mobss['EXPECTGAIN'] * 100, 2))
        phase4_mobss.insert(9, 'SCHEDULEDTIME', sched)
        sc = phase4_mobss['SCHEDULEDTIME']
        phase4_mobss['DOWNTIMEINMINS'] = pd.to_numeric((phase4_mobss['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = phase4_mobss['DOWNTIMEINMINS']
        a = sc - dm
        phase4_mobss.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(phase4_mobss['UPTIMEPERCENTAGE'] * 100), 2)
        phase4_mobss.insert(6, 'REVENUEACHIEVEMENT', round(phase4_mobss['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        phase4_mobss.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        phase4_mobss.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_mobss.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_mobss.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        phase4_mobss.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_mobss.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        phase4_mobss.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        phase4_mobss.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, phase4_mobss)
        get_factorsATT(phase4_mobss, 'AT&TPhase4_MobSS')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')
    elif text == 'NurseLine':
        nurseline = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='NurseLine')
        nurseline.columns = map(lambda x: str(x).upper(), nurseline.columns)
        nurseline.columns = map(lambda x: str(x).replace('_', ''), nurseline.columns)
        nurseline.columns = map(lambda x: str(x).replace(' ', ''), nurseline.columns)
        nurseline.columns = map(lambda x: str(x).rstrip(), nurseline.columns)
        nurseline.insert(3, 'EXPECTGAIN', e)
        nurseline.insert(4, 'GAIN', g)
        summ = nurseline.sum(axis=0)
        nurseline.insert(5, 'GAINACHIEVEMENT', round(nurseline['GAIN'] / nurseline['EXPECTGAIN'] * 100, 2))
        nurseline.insert(9, 'SCHEDULEDTIME', sched)
        sc = nurseline['SCHEDULEDTIME']
        nurseline['DOWNTIMEINMINS'] = pd.to_numeric((nurseline['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = nurseline['DOWNTIMEINMINS']
        a = sc - dm
        nurseline.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(nurseline['UPTIMEPERCENTAGE'] * 100), 2)
        nurseline.insert(6, 'REVENUEACHIEVEMENT', round(nurseline['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        nurseline.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        nurseline.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        nurseline.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        nurseline.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        nurseline.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        nurseline.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        nurseline.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        nurseline.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, nurseline)
        get_factorsNurseLine(nurseline, 'NurseLine')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')
    elif text == 'HouseCalls':
        housecalls = pd.read_excel('test.xlsx', index_col=0, header=0, sheet_name='HouseCalls')
        housecalls.columns = map(lambda x: str(x).upper(), housecalls.columns)
        housecalls.columns = map(lambda x: str(x).replace('_', ''), housecalls.columns)
        housecalls.columns = map(lambda x: str(x).replace(' ', ''), housecalls.columns)
        housecalls.columns = map(lambda x: str(x).rstrip(), housecalls.columns)
        housecalls.insert(3, 'EXPECTGAIN', e)
        housecalls.insert(4, 'GAIN', g)
        summ = housecalls.sum(axis=0)
        housecalls.insert(5, 'GAINACHIEVEMENT', round(housecalls['GAIN'] / housecalls['EXPECTGAIN'] * 100, 2))
        housecalls.insert(9, 'SCHEDULEDTIME', sched)
        sc = housecalls['SCHEDULEDTIME']
        housecalls['DOWNTIMEINMINS'] = pd.to_numeric((housecalls['DOWNTIMEINMINS']), errors='coerce').fillna(0)
        dm = housecalls['DOWNTIMEINMINS']
        a = sc - dm
        housecalls.insert(10, 'UPTIMEPERCENTAGE', (a / sc).tolist())
        EBP_UPTIME = round(mean(housecalls['UPTIMEPERCENTAGE'] * 100), 2)
        housecalls.insert(6, 'REVENUEACHIEVEMENT', round(housecalls['GAINACHIEVEMENT'] * EBP_UPTIME / 100, 2))
        housecalls.insert(12, 'FAILEDROUTEIMPACT', summ['FAILROUTEPERC'] * 100 / summ['REVENUEIMPACT'])
        housecalls.insert(13, 'ONAGENTSLAIMPACT', summ['ONAGENTSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        housecalls.insert(14, 'ONCALLSLAIMPACT', summ['ONCALLSLA%AGE'] * 100 / summ['REVENUEIMPACT'])
        housecalls.insert(15, '1-1CALLSIMPACT', summ['1-1CALLS%AGEWITHOUTSLABLOWNS'] * 100 / summ['REVENUEIMPACT'])
        housecalls.insert(16, 'ONEVALUATIONERRIMPACT', summ['OFFEVALUATIONERRCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        housecalls.insert(17, 'LOOKUPFAILUREIMPACT', summ['LOOKUPFAILUREPERC'] * 100 / summ['REVENUEIMPACT'])
        housecalls.insert(18, 'UNKNOWNAGENTIMPACT', summ['UNKNOWNAGENTCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        housecalls.insert(19, 'CGNOTFOUNDIMPACT', summ['CGNOTFOUNDCALLS%AGE'] * 100 / summ['REVENUEIMPACT'])
        write_excel('Test.xlsx', text, housecalls)
        get_factorsHouseCalls(housecalls, 'HouseCalls')
        messagebox.showinfo('Data Saved', 'Calculations Added to File')


def runFile(filename):
    date = 0
    account = 'abc'
    wb = load_workbook(filename)
    wb2 = load_workbook('Test.xlsx')
    sheet1 = wb['BYGS_Sales']
    sheet2 = wb['BYGS_Care']
    sheet3 = wb['Sky Platform Classic']
    sheet4 = wb['VM Care']
    sheet5 = wb['LM']
    sheet6 = wb['Verizon']
    sheet7 = wb['Nurseline']
    sheet8 = wb['HouseCalls']
    sheet9 = wb['VM Cable Care']
    sheet28 = wb['SkyService']
    sheet18 = wb['AT&T Phase 3']
    sheet22 = wb['AT&T Phase4']
    sheet10 = wb2['Bygs_Sales']
    sheet11 = wb2['Bygs_Care']
    sheet29 = wb2['SkyService']
    sheet12 = wb2['SKY_PlatformClassic']
    sheet13 = wb2['VMCare']
    sheet30 = wb2['VM Cable Care']
    sheet14 = wb2['LM']
    sheet15 = wb2['Verizon']
    sheet16 = wb2['NurseLine']
    sheet17 = wb2['HouseCalls']
    sheet19 = wb2['AT&TPhase3_GenHSCLG']
    sheet20 = wb2['AT&TPhase3_GENHSSALES']
    sheet21 = wb2['AT&TPhase3_GenDTVSales']
    sheet23 = wb2['AT&TPhase4_MobCLG']
    sheet24 = wb2['AT&TPhase4_ISMCLG']
    sheet25 = wb2['AT&TPhase4_MobSS']
    sheet26 = wb2['AT&TPhase4_DMDR']
    sheet27 = wb2['AT&TPhase4_ISMSVC']
    for i in range(1, sheet1.max_row + 1):
        for j in range(1, sheet1.max_column + 1):
            sheet10.cell(row=i, column=j).value = sheet1.cell(row=i, column=j).value

    else:
        for i in range(1, sheet2.max_row + 1):
            for j in range(1, sheet2.max_column + 1):
                sheet11.cell(row=i, column=j).value = sheet2.cell(row=i, column=j).value

        else:
            for i in range(1, sheet28.max_row + 1):
                for j in range(1, sheet28.max_column + 1):
                    sheet29.cell(row=i, column=j).value = sheet28.cell(row=i, column=j).value

            else:
                for i in range(1, sheet3.max_row + 1):
                    for j in range(1, sheet3.max_column + 1):
                        sheet12.cell(row=i, column=j).value = sheet3.cell(row=i, column=j).value

                else:
                    for i in range(1, sheet4.max_row + 1):
                        for j in range(1, sheet4.max_column + 1):
                            sheet13.cell(row=i, column=j).value = sheet4.cell(row=i, column=j).value

                    else:
                        for i in range(1, sheet9.max_row + 1):
                            for j in range(1, sheet9.max_column + 1):
                                sheet30.cell(row=i, column=j).value = sheet9.cell(row=i, column=j).value

                        else:
                            for i in range(1, sheet5.max_row + 1):
                                for j in range(1, sheet5.max_column + 1):
                                    sheet14.cell(row=i, column=j).value = sheet5.cell(row=i, column=j).value

                            else:
                                for i in range(1, sheet6.max_row + 1):
                                    for j in range(1, sheet6.max_column + 1):
                                        sheet15.cell(row=i, column=j).value = sheet6.cell(row=i, column=j).value

                                else:
                                    for i in range(1, sheet7.max_row + 1):
                                        for j in range(1, sheet7.max_column + 1):
                                            sheet16.cell(row=i, column=j).value = sheet7.cell(row=i, column=j).value

                                    else:
                                        for i in range(1, sheet8.max_row + 1):
                                            for j in range(1, sheet8.max_column + 1):
                                                sheet17.cell(row=i, column=j).value = sheet8.cell(row=i, column=j).value

                                        else:
                                            for i in range(1, sheet18.max_row + 1):
                                                for j in range(1, sheet18.max_column + 1):
                                                    if i == 1:
                                                        sheet19.cell(row=i, column=j).value = sheet18.cell(row=i, column=j).value
                                                        sheet20.cell(row=i, column=j).value = sheet18.cell(row=i, column=j).value
                                                        sheet21.cell(row=i, column=j).value = sheet18.cell(row=i, column=j).value
                                                    else:
                                                        if sheet18.cell(row=i, column=2).value == 'attgenhsclg':
                                                            sheet19.cell(row=i, column=j).value = sheet18.cell(row=i, column=j).value
                                                            date = sheet19.cell(row=i, column=3).value
                                                            account = sheet19.cell(row=i, column=1).value
                                                        else:
                                                            if sheet18.cell(row=i, column=2).value == 'ATTGENHSSALES':
                                                                sheet20.cell(row=i, column=j).value = sheet18.cell(row=i, column=j).value
                                                                sheet20.cell(row=i, column=3).value = date
                                                                sheet20.cell(row=i, column=1).value = account
                                                            else:
                                                                if sheet18.cell(row=i, column=2).value == 'attgendtvsales':
                                                                    sheet21.cell(row=i, column=j).value = sheet18.cell(row=i, column=j).value
                                                                    sheet21.cell(row=i, column=3).value = date
                                                                    sheet21.cell(row=i, column=1).value = account

                                            else:
                                                for i in range(1, sheet22.max_row + 1):
                                                    for j in range(1, sheet22.max_column + 1):
                                                        if i == 1:
                                                            sheet23.cell(row=i, column=j).value = sheet22.cell(row=i, column=j).value
                                                            sheet24.cell(row=i, column=j).value = sheet22.cell(row=i, column=j).value
                                                            sheet25.cell(row=i, column=j).value = sheet22.cell(row=i, column=j).value
                                                            sheet26.cell(row=i, column=j).value = sheet22.cell(row=i, column=j).value
                                                            sheet27.cell(row=i, column=j).value = sheet22.cell(row=i, column=j).value
                                                        else:
                                                            if sheet22.cell(row=i, column=2).value == 'MobCLG':
                                                                sheet23.cell(row=i, column=j).value = sheet22.cell(row=i, column=j).value
                                                                date = sheet23.cell(row=i, column=3).value
                                                                account = sheet23.cell(row=i, column=1).value
                                                            else:
                                                                if sheet22.cell(row=i, column=2).value == 'IsmCLG':
                                                                    sheet24.cell(row=i, column=j).value = sheet22.cell(row=i, column=j).value
                                                                    sheet24.cell(row=i, column=3).value = date
                                                                    sheet24.cell(row=i, column=1).value = account
                                                                else:
                                                                    if sheet22.cell(row=i, column=2).value == 'Mobss':
                                                                        sheet25.cell(row=i, column=j).value = sheet22.cell(row=i, column=j).value
                                                                        sheet25.cell(row=i, column=3).value = date
                                                                        sheet25.cell(row=i, column=1).value = account
                                                                    else:
                                                                        if sheet22.cell(row=i, column=2).value == 'Dmdr':
                                                                            sheet26.cell(row=i, column=j).value = sheet22.cell(row=i, column=j).value
                                                                            sheet26.cell(row=i, column=3).value = date
                                                                            sheet26.cell(row=i, column=1).value = account
                                                                        else:
                                                                            if sheet22.cell(row=i, column=2).value == 'ISMSVC':
                                                                                sheet27.cell(row=i, column=j).value = sheet22.cell(row=i, column=j).value
                                                                                sheet27.cell(row=i, column=3).value = date
                                                                                sheet27.cell(row=i, column=1).value = account

                                                else:
                                                    index_row = []
                                                    rows_to_delete = [
                                                     None, '', ' ']
                                                    for i in wb2.sheetnames:
                                                        ws = wb2[i]
                                                        if not i == 'AT&TPhase3_GenDTVSales':
                                                            if not i == 'AT&TPhase3_GenHSCLG':
                                                                if not i == 'AT&TPhase3_GENHSSALES':
                                                                    if not i == 'AT&TPhase4_DMDR':
                                                                        if not i == 'AT&TPhase4_ISMCLG':
                                                                            if not i == 'AT&TPhase4_ISMSVC':
                                                                                if not i == 'AT&TPhase4_MobCLG':
                                                                                    if not i == 'AT&TPhase4_MobSS':
                                                                                        if i == 'LM':
                                                                                            pass
                                                        column_b = range(1, ws.max_row)
                                                        for i in reversed(column_b):
                                                            if ws.cell(i, 2).value in rows_to_delete:
                                                                ws.delete_rows(ws.cell(i, 2).row)

                                                    else:
                                                        wb.save(filename)
                                                        wb2.save('Test.xlsx')


def SchemaBuilding():
    Final_File = pd.concat([sales, care, sky, skyservice, vmcare, vmcablecare, LM, Verizon, phase3_gendtvsales, phase3_gendhsclg, phase3_gendhssales, phase4_dmdr, phase4_ismclg, phase4_ismsvc, phase4_mobclg, phase4_mobss, nurseline, housecalls])
    Final_File.to_excel('Schema.xlsx', index=False)


class NewWindow(Toplevel):

    def __init__(self, master=None):
        super().__init__(master=master)
        self.title('High Value Accounts')
        self.geometry('1000x1000')
        load = Image.open('Background.jpg')
        bg = ImageTk.PhotoImage(load)
        background_label = Label(self, image=bg)
        background_label.place(x=0, y=0, relwidth=1, relheight=1)
        l11 = Label(self, text='')
        l11.pack()
        l12 = Label(self, text='')
        l12.pack()
        l13 = Label(self, text='')
        l13.pack()
        l14 = Label(self, text='')
        l14.pack()
        l15 = Label(self, text='')
        l15.pack()
        l16 = Label(self, text='')
        l16.pack()
        l17 = Label(self, text='')
        l17.pack()
        self.sales = Button(self, text='Bouygues Sales', padx=10)
        self.sales.bind('<Button>', lambda e: NewWindow1(master, 'Bygs_Sales'))
        self.sales.pack()
        self.care = Button(self, text='Bouygues Care', padx=10)
        self.care.bind('<Button>', lambda e: NewWindow1(master, 'Bygs_Care'))
        self.care.pack()
        self.sky = Button(self, text='Sky Platform Classic', padx=10)
        self.sky.bind('<Button>', lambda e: NewWindow1(master, 'SKY_PlatformClassic'))
        self.sky.pack()
        self.skyservice = Button(self, text='Sky Service', padx=10)
        self.skyservice.bind('<Button>', lambda e: NewWindow1(master, 'SkyService'))
        self.skyservice.pack()
        self.verizon = Button(self, text='Verizon Wireless', padx=10)
        self.verizon.bind('<Button>', lambda e: NewWindow1(master, 'Verizon'))
        self.verizon.pack()
        self.LM = Button(self, text='LM', padx=10)
        self.LM.bind('<Button>', lambda e: NewWindow1(master, 'LM'))
        self.LM.pack()
        self.VMCare = Button(self, text='Virgin Media Care', padx=10)
        self.VMCare.bind('<Button>', lambda e: NewWindow1(master, 'VMCare'))
        self.VMCare.pack()
        self.VMCableCare = Button(self, text='Virgin Media Cable Care', padx=10)
        self.VMCableCare.bind('<Button>', lambda e: NewWindow1(master, 'VM Cable Care'))
        self.VMCableCare.pack()
        self.phase3_gendtvsales = Button(self, text='AT&T Phase3 GenDTV Sales', padx=10)
        self.phase3_gendtvsales.bind('<Button>', lambda e: NewWindow1(master, 'AT&TPhase3_GenDTVSales'))
        self.phase3_gendtvsales.pack()
        self.phase3_gendhsclg = Button(self, text='AT&T Phase3 GenHSCLG', padx=10)
        self.phase3_gendhsclg.bind('<Button>', lambda e: NewWindow1(master, 'AT&TPhase3_GenHSCLG'))
        self.phase3_gendhsclg.pack()
        self.phase3_gendhssales = Button(self, text='AT&T Phase3 GENHS SALES', padx=10)
        self.phase3_gendhssales.bind('<Button>', lambda e: NewWindow1(master, 'AT&TPhase3_GENHSSALES'))
        self.phase3_gendhssales.pack()
        self.phase4_dmdr = Button(self, text='AT&T Phase4 DMDR', padx=10)
        self.phase4_dmdr.bind('<Button>', lambda e: NewWindow1(master, 'AT&TPhase4_DMDR'))
        self.phase4_dmdr.pack()
        self.phase4_ismclg = Button(self, text='AT&T Phase4 ISMCLG', padx=10)
        self.phase4_ismclg.bind('<Button>', lambda e: NewWindow1(master, 'AT&TPhase4_ISMCLG'))
        self.phase4_ismclg.pack()
        self.phase4_ismsvc = Button(self, text='AT&T Phase4 ISMSVC', padx=10)
        self.phase4_ismsvc.bind('<Button>', lambda e: NewWindow1(master, 'AT&TPhase4_ISMSVC'))
        self.phase4_ismsvc.pack()
        self.phase4_mobclg = Button(self, text='AT&T Phase4 MobCLG', padx=10)
        self.phase4_mobclg.bind('<Button>', lambda e: NewWindow1(master, 'AT&TPhase4_MobCLG'))
        self.phase4_mobclg.pack()
        self.phase4_mobss = Button(self, text='AT&T Phase4 MobSS', padx=10)
        self.phase4_mobss.bind('<Button>', lambda e: NewWindow1(master, 'AT&TPhase4_MobSS'))
        self.phase4_mobss.pack()
        self.NurseLine = Button(self, text='UHG NurseLine', padx=10)
        self.NurseLine.bind('<Button>', lambda e: NewWindow1(master, 'NurseLine'))
        self.NurseLine.pack()
        self.HouseCalls = Button(self, text='UHG House Calls', padx=10)
        self.HouseCalls.bind('<Button>', lambda e: NewWindow1(master, 'HouseCalls'))
        self.HouseCalls.pack()
        self.schemabuilding = Button(self, text='Centralized Schema', padx=10, pady=10, command=SchemaBuilding)
        self.schemabuilding.pack()
        self.mainloop()


class NewWindow1(Toplevel):

    def __init__(self, master, text):
        super().__init__(master=master)
        load = Image.open('EMEA_ROW.gif')
        bg = ImageTk.PhotoImage(load)
        background_label = Label(self, image=bg)
        background_label.place(x=0, y=0, relwidth=1, relheight=1)
        l11 = Label(self, text='')
        l11.pack()
        l12 = Label(self, text='')
        l12.pack()
        l16 = Label(self, text='')
        l16.pack()
        l17 = Label(self, text='')
        l17.pack()
        l1 = Label(self, text='Expect Gain')
        l1.pack()
        self.title('Enter Records')
        self.geometry('500x500')
        self.e = Entry(self)
        self.e.pack()
        self.e.focus_set()
        l2 = Label(self, text='Gain')
        l2.pack()
        self.e1 = Entry(self)
        self.e1.pack()
        self.e1.focus_set()
        l3 = Label(self, text='Scheduled Time')
        l3.pack()
        self.e2 = Entry(self)
        self.e2.pack()
        self.e2.focus_set()
        self.b = Button(self, text='Save')
        self.b.bind('<Button>', lambda e: insert_record(text, self.e.get(), self.e1.get(), self.e2.get()))
        self.b.pack(side='bottom', pady=10)
        self.mainloop()


def copydata():
    global e
    string = e.get()
    runFile(string)
    messagebox.showinfo('Message', 'Data Replicated from Daily Report')
    NewWindow(root)


root = Tk()
root.title('Please Enter File Name To Load')
e = Entry(root, width=100)
e.pack()
e.focus_set()
b = Button(root, text='Load File', command=copydata)
b.pack(side='bottom', pady=10)
root.mainloop()
# global wb ## Warning: Unused global